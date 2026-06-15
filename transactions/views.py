from io import BytesIO
from django.http import HttpResponse
from django.db.models import Sum
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from .models import Transaction
from .serializers import TransactionSerializer
import openpyxl, datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from budgets.models import Budget

MOIS_FR = ['Jan','Fév','Mar','Avr','Mai','Jui','Jul','Aoû','Sep','Oct','Nov','Déc']


class TransactionListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        transactions = Transaction.objects.filter(
            id_entreprise=request.user.id_entreprise
        ).select_related('id_categorie', 'id_utilisateur').order_by('-date_transaction')
        serializer = TransactionSerializer(transactions, many=True)
        return Response(serializer.data)

    def post(self, request):
        data = request.data.copy()
        data['id_entreprise']  = request.user.id_entreprise.id
        data['id_utilisateur'] = request.user.id
        serializer = TransactionSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class TransactionDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        try:
            transaction = Transaction.objects.get(pk=pk, id_entreprise=request.user.id_entreprise)
            serializer  = TransactionSerializer(transaction, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Transaction.DoesNotExist:
            return Response({'error': 'Transaction non trouvée'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        try:
            transaction = Transaction.objects.get(pk=pk, id_entreprise=request.user.id_entreprise)
            transaction.delete()
            return Response({'message': 'Transaction supprimée !'}, status=status.HTTP_204_NO_CONTENT)
        except Transaction.DoesNotExist:
            return Response({'error': 'Transaction non trouvée'}, status=status.HTTP_404_NOT_FOUND)


class DashboardStatsView(APIView):
    """Retourne tous les agrégats nécessaires au tableau de bord en un seul appel."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from alertes.models import Alerte
        from dettes_factures.models import DetteFacture

        entreprise = request.user.id_entreprise
        today      = datetime.date.today()
        mois_debut = datetime.date(today.year, today.month, 1)

        def agg(type_, **kwargs):
            return float(Transaction.objects.filter(
                id_entreprise=entreprise, type=type_, **kwargs
            ).aggregate(t=Sum('montant'))['t'] or 0)

        entrees_total = agg('Entree')
        sorties_total = agg('Sortie')
        entrees_mois  = agg('Entree', date_transaction__gte=mois_debut)
        sorties_mois  = agg('Sortie', date_transaction__gte=mois_debut)

        nb_transactions   = Transaction.objects.filter(id_entreprise=entreprise).count()
        alertes_non_lues  = Alerte.objects.filter(id_entreprise=entreprise, statut='non_lue').count()
        dettes_en_retard  = DetteFacture.objects.filter(id_entreprise=entreprise, statut='en_retard').count()

        # Évolution sur les 6 derniers mois
        evolution = []
        for i in range(5, -1, -1):
            y, m = today.year, today.month - i
            while m <= 0:
                m += 12; y -= 1
            debut = datetime.date(y, m, 1)
            fin   = datetime.date(y + 1, 1, 1) if m == 12 else datetime.date(y, m + 1, 1)
            evolution.append({
                'mois':     MOIS_FR[m - 1],
                'revenus':  agg('Entree', date_transaction__gte=debut, date_transaction__lt=fin),
                'depenses': agg('Sortie', date_transaction__gte=debut, date_transaction__lt=fin),
            })

        # 5 dernières transactions pour le tableau de bord
        recentes = Transaction.objects.filter(id_entreprise=entreprise).select_related(
            'id_categorie'
        ).order_by('-date_transaction')[:5]
        recentes_data = [{
            'id':               t.id,
            'libelle':          t.libelle,
            'type':             t.type,
            'montant':          str(t.montant),
            'date_transaction': str(t.date_transaction),
            'statut':           t.statut,
            'categorie':        t.id_categorie.nom_categorie if t.id_categorie else '',
        } for t in recentes]

        # ── Score de santé financière (0-100) ──────────────────────────────
        criteres = []

        # 1. Flux du mois — entrées vs sorties (30 pts)
        if sorties_mois > 0:
            ratio_mois = entrees_mois / sorties_mois
            if ratio_mois >= 1.5:   pts1 = 30
            elif ratio_mois >= 1.2: pts1 = 20
            elif ratio_mois >= 1.0: pts1 = 10
            else:                   pts1 = 0
            detail1 = f'Ratio {ratio_mois:.1f}×'
        elif entrees_mois > 0:
            pts1 = 30; detail1 = 'Aucune dépense ce mois'
        else:
            pts1 = 0; detail1 = 'Aucune activité ce mois'
        criteres.append({'label': 'Flux du mois', 'pts': pts1, 'max': 30, 'detail': detail1})

        # 2. Solde global positif (20 pts)
        solde_val = entrees_total - sorties_total
        pts2 = 20 if solde_val > 0 else 0
        criteres.append({'label': 'Solde positif', 'pts': pts2, 'max': 20,
                         'detail': f'{int(solde_val):,} XOF'.replace(',', ' ')})

        # 3. Dettes en retard (20 pts)
        if dettes_en_retard == 0:   pts3 = 20
        elif dettes_en_retard <= 2: pts3 = 10
        else:                       pts3 = 0
        criteres.append({'label': 'Dettes à jour', 'pts': pts3, 'max': 20,
                         'detail': f'{dettes_en_retard} en retard' if dettes_en_retard else 'Aucun retard'})

        # 4. Budgets respectés (20 pts)
        budgets_qs = Budget.objects.filter(id_entreprise=entreprise)
        total_b = budgets_qs.count()
        if total_b > 0:
            depasses = sum(1 for b in budgets_qs if b.is_depasse())
            pts4 = round(((total_b - depasses) / total_b) * 20)
            detail4 = f'{total_b - depasses}/{total_b} respecté(s)'
        else:
            pts4 = 20; detail4 = 'Aucun budget défini'
        criteres.append({'label': 'Budgets', 'pts': pts4, 'max': 20, 'detail': detail4})

        # 5. Activité régulière ce mois (10 pts)
        nb_tx_mois = Transaction.objects.filter(
            id_entreprise=entreprise, date_transaction__gte=mois_debut
        ).count()
        if nb_tx_mois >= 10:   pts5 = 10
        elif nb_tx_mois >= 3:  pts5 = 7
        elif nb_tx_mois >= 1:  pts5 = 4
        else:                  pts5 = 0
        criteres.append({'label': 'Activité', 'pts': pts5, 'max': 10,
                         'detail': f'{nb_tx_mois} transaction(s) ce mois'})

        score_sante = pts1 + pts2 + pts3 + pts4 + pts5
        if score_sante >= 80:   niveau = 'Excellente santé'
        elif score_sante >= 60: niveau = 'Bonne santé'
        elif score_sante >= 40: niveau = 'À surveiller'
        else:                   niveau = 'Attention requise'

        # ── Prévision de trésorerie à 30 jours ────────────────────────────
        j90_debut = today - datetime.timedelta(days=90)
        j30_fin   = today + datetime.timedelta(days=30)

        entrees_90j = agg('Entree', date_transaction__gte=j90_debut)
        sorties_90j = agg('Sortie', date_transaction__gte=j90_debut)
        flux_journalier = (entrees_90j - sorties_90j) / 90
        solde_actuel    = entrees_total - sorties_total
        solde_prevu_30j = solde_actuel + flux_journalier * 30

        # Dettes arrivant à échéance dans les 30 prochains jours
        dettes_30j = DetteFacture.objects.filter(
            id_entreprise=entreprise,
            statut__in=['en_cours', 'partiellement_paye', 'en_retard'],
            date_echeance__lte=j30_fin,
        )
        a_recevoir = sum(
            float(d.montant_total) - float(d.montant_paye or 0)
            for d in dettes_30j if d.type == 'Client'
        )
        a_payer = sum(
            float(d.montant_total) - float(d.montant_paye or 0)
            for d in dettes_30j if d.type == 'Fournisseur'
        )
        solde_prevu_ajuste = solde_prevu_30j + a_recevoir - a_payer

        if flux_journalier > 500:      tendance = 'hausse'
        elif flux_journalier < -500:   tendance = 'baisse'
        else:                          tendance = 'stable'

        prevision = {
            'solde_actuel':         round(solde_actuel),
            'solde_prevu_30j':      round(solde_prevu_ajuste),
            'flux_journalier':      round(flux_journalier),
            'a_recevoir_30j':       round(a_recevoir),
            'a_payer_30j':          round(a_payer),
            'tendance':             tendance,
            'nb_dettes_echeance':   dettes_30j.count(),
        }

        return Response({
            'solde':             entrees_total - sorties_total,
            'entrees_total':     entrees_total,
            'sorties_total':     sorties_total,
            'entrees_mois':      entrees_mois,
            'sorties_mois':      sorties_mois,
            'nb_transactions':   nb_transactions,
            'alertes_non_lues':  alertes_non_lues,
            'dettes_en_retard':  dettes_en_retard,
            'evolution':         evolution,
            'transactions_recentes': recentes_data,
            'score_sante':       score_sante,
            'score_niveau':      niveau,
            'score_criteres':    criteres,
            'prevision':         prevision,
        })


class TransactionExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Transaction.objects.filter(
            id_entreprise=request.user.id_entreprise
        ).select_related('id_categorie', 'id_utilisateur').order_by('-date_transaction')

        if request.query_params.get('type'):
            qs = qs.filter(type=request.query_params['type'])
        if request.query_params.get('statut'):
            qs = qs.filter(statut=request.query_params['statut'])
        if request.query_params.get('date_debut'):
            qs = qs.filter(date_transaction__gte=request.query_params['date_debut'])
        if request.query_params.get('date_fin'):
            qs = qs.filter(date_transaction__lte=request.query_params['date_fin'])
        if request.query_params.get('categorie'):
            qs = qs.filter(id_categorie=request.query_params['categorie'])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transactions'

        header_fill = PatternFill('solid', fgColor='0F172A')
        header_font = Font(bold=True, color='22C55E', size=11)
        entree_font = Font(color='22C55E', bold=True)
        sortie_font = Font(color='EF4444', bold=True)

        headers = ['#', 'Libellé', 'Type', 'Montant (XOF)', 'Catégorie', 'Date', 'Statut', 'Saisi par']
        widths  = [5,   35,        10,     18,              20,           14,     14,        20]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = w

        statuts = {'validee': 'Validée', 'en_attente': 'En attente'}
        for row, tx in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=tx.id)
            ws.cell(row=row, column=2, value=tx.libelle or '')
            tc = ws.cell(row=row, column=3, value='Entrée' if tx.type == 'Entree' else 'Sortie')
            tc.font = entree_font if tx.type == 'Entree' else sortie_font
            mc = ws.cell(row=row, column=4, value=float(tx.montant))
            mc.font = entree_font if tx.type == 'Entree' else sortie_font
            mc.number_format = '#,##0'
            ws.cell(row=row, column=5, value=tx.id_categorie.nom_categorie if tx.id_categorie else '')
            ws.cell(row=row, column=6, value=str(tx.date_transaction))
            ws.cell(row=row, column=7, value=statuts.get(tx.statut, tx.statut))
            ws.cell(row=row, column=8, value=tx.id_utilisateur.username if tx.id_utilisateur else '')

        ws.freeze_panes = 'A2'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nom = request.user.id_entreprise.nom if request.user.id_entreprise else 'export'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="transactions_{nom}.xlsx"'
        return response
